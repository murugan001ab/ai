import cv2
import numpy as np
from ultralytics import YOLO
import time
import os
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO

# ========================= CONFIG =========================
model = YOLO('yolo26n-pose.pt') 
IDLE_THRESHOLD_SECONDS = 5
MOTION_SENSITIVITY = 0.005 
EXCEL_OUTPUT = "Inactivity_Log.xlsx"
ZONE_ID = "Zone 1"

# Initialize Background Subtractor
fgbg = cv2.createBackgroundSubtractorMOG2(history=500, varThreshold=50, detectShadows=False)

# Tracking State
person_history = {}                
idle_captured_flag = defaultdict(bool) 
last_idle_crop = {}

# --- ID STABILIZATION CONFIG ---
id_map = {}          # Dictionary to link YOLO IDs to Sequential IDs
next_avail_id = 1    # Counter for sequential IDs

# Initialize Excel
if not os.path.exists(EXCEL_OUTPUT):
    wb = Workbook(); ws = wb.active; ws.title = "Inactivity Log"
    ws.append(["Timestamp", "Zone ID", "Worker ID", "Total Idle Time", "Image"])
    wb.save(EXCEL_OUTPUT)

def log_to_excel(timestamp, zone, display_id, duration, frame_crop):
    if frame_crop is None: return
    try:
        wb = load_workbook(EXCEL_OUTPUT); ws = wb.active; next_row = ws.max_row + 1
        is_success, buffer = cv2.imencode(".jpg", frame_crop)
        if is_success:
            img_buffer = BytesIO(buffer)
            xl_img = XLImage(img_buffer); xl_img.width, xl_img.height = 150, 100
            ws.append([timestamp, zone, f"Worker {display_id}", f"{int(duration)}s", "Logged"])
            ws.add_image(xl_img, f'E{next_row}')
            ws.row_dimensions[next_row].height = 80
            wb.save(EXCEL_OUTPUT)
    except Exception as e: print(f"Excel Log Error: {e}")

print("Monitoring... Press 'q' to stop.")
results_generator = model.track(source="1", persist=True, tracker="bytetrack.yaml", stream=True)

for result in results_generator:
    current_time = time.time()
    frame = result.orig_img
    
    fg_mask = fgbg.apply(frame)
    _, fg_mask = cv2.threshold(fg_mask, 200, 255, cv2.THRESH_BINARY)

    if result.boxes.id is not None:
        boxes = result.boxes.xyxy.cpu().numpy()
        track_ids = result.boxes.id.cpu().numpy().astype(int)
        
        for i, (box, orig_track_id) in enumerate(zip(boxes, track_ids)):
            x1, y1, x2, y2 = map(int, box)
            
            # --- ID STABILIZATION LOGIC ---
            if orig_track_id not in id_map:
                id_map[orig_track_id] = next_avail_id
                next_avail_id += 1
            display_id = id_map[orig_track_id] 

            # --- MICRO-MOVEMENT DETECTION ---
            roi_motion = fg_mask[max(0, y1):y2, max(0, x1):x2]
            motion_score = np.count_nonzero(roi_motion) / roi_motion.size if roi_motion.size > 0 else 0

            if orig_track_id not in person_history:
                person_history[orig_track_id] = {'start_time': current_time}

            if motion_score > MOTION_SENSITIVITY:
                elapsed = current_time - person_history[orig_track_id]['start_time']
                if idle_captured_flag[orig_track_id] and elapsed >= IDLE_THRESHOLD_SECONDS:
                    ts = time.strftime("%Y-%m-%d %H:%M:%S")
                    log_to_excel(ts, ZONE_ID, display_id, elapsed, last_idle_crop.get(orig_track_id))
                
                person_history[orig_track_id] = {'start_time': current_time}
                idle_captured_flag[orig_track_id] = False
                status, color = "Active", (0, 255, 0)
            else:
                elapsed = current_time - person_history[orig_track_id]['start_time']
                if elapsed >= IDLE_THRESHOLD_SECONDS:
                    status, color = f"IDLE {int(elapsed)}s", (0, 0, 255)
                    if not idle_captured_flag[orig_track_id]:
                        last_idle_crop[orig_track_id] = frame[max(0, y1-10):y2+10, max(0, x1-10):x2+10].copy()
                        idle_captured_flag[orig_track_id] = True
                else:
                    status, color = "Watching...", (0, 255, 255)

            # Visuals
            cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
            cv2.putText(frame, f"Worker {display_id}: {status}", (x1, y1 - 10), 
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    cv2.imshow("AI Micro-Motion Monitor", frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cv2.destroyAllWindows()

